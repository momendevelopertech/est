from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import openpyxl
import xlrd
from docx import Document


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_phone(value: Any) -> str:
    phone = normalize_text(value)
    if not phone:
        return ""
    compact = re.sub(r"[\s\-().]", "", phone)
    if compact.startswith("+"):
        return f"+{compact[1:].replace('+', '')}"
    return compact.replace("+", "")


def normalize_email(value: Any) -> str:
    return normalize_text(value).lower()


def normalize_identifier(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_text(value))


def clean_room_name(value: Any) -> str:
    room = normalize_text(value)
    if not room:
        return ""
    room = room.replace("*", "")
    room = re.sub(r"(?i)\s*lecture hall\s*", "", room)
    room = room.replace(".0", "")
    room = re.sub(r"\s+", " ", room)
    return room.strip()


def parse_integer(value: Any) -> int | None:
    raw = normalize_text(value)
    if not raw:
        return None
    lowered = raw.lower()
    if lowered in {"#n/a", "na", "none"}:
        return None
    digits = raw.replace(",", "")
    try:
        if "." in digits:
            return int(float(digits))
        return int(digits)
    except ValueError:
        return None


def maybe_append_person(rows: list[dict[str, Any]], payload: dict[str, Any]) -> None:
    name_en = normalize_text(payload.get("nameEn"))
    if not name_en:
        return

    lowered = name_en.lower()
    if lowered in {
        "full english name (at least 4 names)",
        "name",
        "name in arabic",
        "list with insurance number\nfue (staff)   january 2026",
    }:
        return

    rows.append(
        {
            "sourceFile": payload.get("sourceFile", ""),
            "sourceSheet": payload.get("sourceSheet", ""),
            "nameEn": name_en,
            "nameAr": normalize_text(payload.get("nameAr")),
            "email": normalize_email(payload.get("email")),
            "phone": normalize_phone(payload.get("phone")),
            "nationalId": normalize_identifier(payload.get("nationalId")),
            "insuranceNumber": normalize_identifier(payload.get("insuranceNumber")),
            "organization": normalize_text(payload.get("organization")),
            "roleLabel": normalize_text(payload.get("roleLabel")),
            "governorate": normalize_text(payload.get("governorate")),
            "preferredCenter": normalize_text(payload.get("preferredCenter")),
            "building": normalize_text(payload.get("building")),
            "location": normalize_text(payload.get("location")),
            "typeLabel": normalize_text(payload.get("typeLabel")),
            "division": normalize_text(payload.get("division")),
        }
    )


def maybe_append_location(rows: list[dict[str, Any]], payload: dict[str, Any]) -> None:
    room = clean_room_name(payload.get("room"))
    building = normalize_text(payload.get("building"))
    if not room or not building:
        return

    rows.append(
        {
            "sourceFile": payload.get("sourceFile", ""),
            "sourceSheet": payload.get("sourceSheet", ""),
            "governorate": normalize_text(payload.get("governorate")),
            "university": normalize_text(payload.get("university")),
            "building": building,
            "floor": normalize_text(payload.get("floor")),
            "room": room,
            "roomType": normalize_text(payload.get("roomType")) or "exam_hall",
            "classCapacity": parse_integer(payload.get("classCapacity")),
            "examCapacity": parse_integer(payload.get("examCapacity")),
            "est1Admitted": parse_integer(payload.get("est1Admitted")),
            "est2Admitted": parse_integer(payload.get("est2Admitted")),
            "isPaperStore": bool(payload.get("isPaperStore", False)),
        }
    )


def parse_aast_docx(
    path: Path,
    location_rows: list[dict[str, Any]],
    governorate: str,
    university: str,
    building_aliases: dict[str, str],
) -> None:
    document = Document(path)

    for table in document.tables:
        rows = []
        for table_row in table.rows:
            values = [normalize_text(cell.text) for cell in table_row.cells]
            if any(values):
                rows.append(values)

        if not rows:
            continue

        building_label = next((cell for cell in reversed(rows[0]) if cell), "")
        canonical_building = building_aliases.get(building_label, building_label)

        if not canonical_building or normalize_text(canonical_building).lower().startswith("total"):
            continue

        for row in rows:
            if len(row) < 7:
                continue

            if not re.fullmatch(r"\d+", normalize_text(row[0])):
                continue

            room_raw = normalize_text(row[2])
            if not room_raw:
                continue

            is_paper_store = "*" in room_raw or any("مخزن" in value for value in row)

            maybe_append_location(
                location_rows,
                {
                    "sourceFile": path.name,
                    "sourceSheet": "docx",
                    "governorate": governorate,
                    "university": university,
                    "building": canonical_building,
                    "floor": normalize_text(row[1]),
                    "room": room_raw,
                    "roomType": "paper_store" if is_paper_store else "exam_hall",
                    "classCapacity": row[3],
                    "examCapacity": row[4],
                    "est1Admitted": row[5],
                    "est2Admitted": row[6],
                    "isPaperStore": is_paper_store,
                },
            )


def parse_sadat_workbook(path: Path, location_rows: list[dict[str, Any]]) -> None:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.active
    current_exam: str | None = None

    faculty_buildings = {
        "physical therapy": "Al Ryada University for Science and Technology - Faculty of Physical Therapy",
        "dentistry": "Al Ryada University for Science and Technology - Faculty of Dentistry",
    }

    for row in sheet.iter_rows(values_only=True):
        values = [normalize_text(cell) for cell in row]
        first = values[0] if values else ""

        if not first:
            continue

        lowered = first.lower()
        if lowered.startswith("est 1"):
            current_exam = "EST1"
            continue
        if lowered.startswith("est 2"):
            current_exam = "EST2"
            continue
        if lowered in {"sadat city", "name", "total"} or not current_exam:
            continue

        faculty = normalize_text(values[1] if len(values) > 1 else "")
        building = faculty_buildings.get(faculty.lower())
        if not building:
            continue

        capacity = parse_integer(values[2] if len(values) > 2 else None)
        maybe_append_location(
            location_rows,
            {
                "sourceFile": path.name,
                "sourceSheet": sheet.title,
                "governorate": "Menoufia",
                "university": "Al Ryada University for Science and Technology",
                "building": building,
                "floor": "",
                "room": first,
                "roomType": "exam_hall",
                "classCapacity": capacity,
                "examCapacity": capacity,
                "est1Admitted": capacity if current_exam == "EST1" else None,
                "est2Admitted": capacity if current_exam == "EST2" else None,
            },
        )


def parse_future_workbook(path: Path, location_rows: list[dict[str, Any]]) -> None:
    workbook = openpyxl.load_workbook(path, data_only=True)

    faculty_buildings = {
        "dentistry": "Future University Faculty of Dentistry",
        "pharmacy": "Future University Faculty of Pharmacy",
        "business": "Future University Faculty of Business",
        "political sc.": "Future University Faculty of Political Science",
        "engineering": "Future University Faculty of Engineering",
        "comp sc.": "Future University Faculty of Computer Science",
    }

    for sheet in workbook.worksheets:
        if sheet.title == "Control_Rooms":
            for row in sheet.iter_rows(min_row=2, values_only=True):
                faculty = normalize_text(row[0] if len(row) > 0 else "")
                room = normalize_text(row[1] if len(row) > 1 else "")
                building = faculty_buildings.get(faculty.lower())
                if not building or not room:
                    continue

                maybe_append_location(
                    location_rows,
                    {
                        "sourceFile": path.name,
                        "sourceSheet": sheet.title,
                        "governorate": "Cairo",
                        "university": "Future University in Egypt",
                        "building": building,
                        "floor": "",
                        "room": room,
                        "roomType": "control_room",
                        "classCapacity": 1,
                        "examCapacity": 1,
                        "est1Admitted": 1,
                        "est2Admitted": 1,
                    },
                )
            continue

        current_exam = "EST1" if "friday" in normalize_text(sheet["A1"].value).lower() else "EST2"
        current_faculty = ""

        for row in sheet.iter_rows(values_only=True):
            values = [normalize_text(cell) for cell in row]
            if len(values) < 2:
                continue

            first = values[0]
            second = values[1]

            if not any(values):
                continue

            if first == "Faculty" or second == "Lecture Hall name:":
                continue

            if second and re.search(r"\bHalls?\b", second, re.IGNORECASE):
                current_faculty = first
                continue

            if second == "Total":
                current_faculty = ""
                continue

            faculty = current_faculty
            room = second

            if first and second and not re.search(r"\bHalls?\b", second, re.IGNORECASE):
                faculty = first
                room = second

            building = faculty_buildings.get(faculty.lower())
            if not building or not room or room == "Total":
                continue

            jan_trial_capacity = parse_integer(values[4] if len(values) > 4 else None)
            exam_capacity = parse_integer(values[3] if len(values) > 3 else None)
            class_capacity = parse_integer(values[2] if len(values) > 2 else None)

            maybe_append_location(
                location_rows,
                {
                    "sourceFile": path.name,
                    "sourceSheet": sheet.title,
                    "governorate": "Cairo",
                    "university": "Future University in Egypt",
                    "building": building,
                    "floor": "",
                    "room": room,
                    "roomType": "exam_hall",
                    "classCapacity": class_capacity,
                    "examCapacity": exam_capacity,
                    "est1Admitted": jan_trial_capacity if current_exam == "EST1" else None,
                    "est2Admitted": jan_trial_capacity if current_exam == "EST2" else None,
                },
            )


def parse_damietta_rooms(path: Path, location_rows: list[dict[str, Any]]) -> None:
    workbook = xlrd.open_workbook(path)

    for sheet_index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_index)
        current_exam = "EST1" if sheet.name.upper() == "EST1" else "EST2"

        for row_index in range(3, sheet.nrows):
            name = normalize_text(sheet.cell_value(row_index, 0))
            faculty = normalize_text(sheet.cell_value(row_index, 1))
            capacity = parse_integer(sheet.cell_value(row_index, 2))

            if not name or name == "Total":
                continue

            faculty_lookup = faculty.lower()
            if "engineering" in faculty_lookup:
                building = "Horus University - Egypt ( HUE ) Faculty of Engineering"
            elif "medecine" in faculty_lookup or "medicine" in faculty_lookup:
                building = "Horus University - Egypt ( HUE ) Faculty of Medicine"
            else:
                continue

            maybe_append_location(
                location_rows,
                {
                    "sourceFile": path.name,
                    "sourceSheet": sheet.name,
                    "governorate": "Damietta",
                    "university": "Horus University - Egypt ( HUE )",
                    "building": building,
                    "floor": "",
                    "room": name,
                    "roomType": "exam_hall",
                    "classCapacity": capacity,
                    "examCapacity": capacity,
                    "est1Admitted": capacity if current_exam == "EST1" else None,
                    "est2Admitted": capacity if current_exam == "EST2" else None,
                },
            )


def parse_ryada_people(path: Path, people_rows: list[dict[str, Any]]) -> None:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = [normalize_text(cell) for cell in row]
        if not any(values):
            continue

        maybe_append_person(
            people_rows,
            {
                "sourceFile": path.name,
                "sourceSheet": sheet.title,
                "nameEn": values[3] if len(values) > 3 else "",
                "nameAr": values[4] if len(values) > 4 else "",
                "email": values[5] if len(values) > 5 else "",
                "phone": values[6] if len(values) > 6 else "",
                "organization": values[7] if len(values) > 7 else "",
                "roleLabel": values[9] if len(values) > 9 else "",
                "insuranceNumber": values[10] if len(values) > 10 else "",
                "nationalId": values[12] if len(values) > 12 else "",
                "governorate": "Menoufia",
                "preferredCenter": values[16] if len(values) > 16 else "",
                "building": values[16] if len(values) > 16 else "",
                "location": "",
                "typeLabel": "EP",
                "division": values[1] if len(values) > 1 else "",
            },
        )


def parse_fue_people(path: Path, people_rows: list[dict[str, Any]]) -> None:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=3, values_only=True):
        values = [normalize_text(cell) for cell in row]
        if not any(values):
            continue

        maybe_append_person(
            people_rows,
            {
                "sourceFile": path.name,
                "sourceSheet": sheet.title,
                "nameEn": values[1] if len(values) > 1 else "",
                "nameAr": "",
                "email": values[2] if len(values) > 2 else "",
                "phone": values[3] if len(values) > 3 else "",
                "organization": values[6] if len(values) > 6 else "",
                "insuranceNumber": values[5] if len(values) > 5 else "",
                "nationalId": values[4] if len(values) > 4 else "",
                "governorate": "Cairo",
                "preferredCenter": values[7] if len(values) > 7 else "",
                "building": "",
                "location": "",
                "typeLabel": "EP",
                "division": "FUE",
            },
        )


def parse_uk_people(path: Path, people_rows: list[dict[str, Any]]) -> None:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.active

    for row in sheet.iter_rows(values_only=True):
        values = [normalize_text(cell) for cell in row]
        if len(values) < 11:
            continue

        maybe_append_person(
            people_rows,
            {
                "sourceFile": path.name,
                "sourceSheet": sheet.title,
                "nameEn": values[2],
                "nameAr": values[3],
                "email": values[4],
                "phone": values[5],
                "organization": values[1],
                "roleLabel": values[6],
                "governorate": values[8],
                "preferredCenter": values[10],
                "building": values[10],
                "location": values[11] if len(values) > 11 else "",
                "typeLabel": values[7],
                "division": values[15] if len(values) > 15 else "",
            },
        )


def parse_damietta_people(path: Path, people_rows: list[dict[str, Any]]) -> None:
    workbook = openpyxl.load_workbook(path, data_only=True)

    layouts = {
        "Sheet2": {"nameEn": 1, "nameAr": 2, "email": 3, "phone": 4, "organization": 5, "insurance": 8, "nationalId": 10, "preferredCenter": 14, "role": 20, "type": 21, "governorate": 22, "building": 24, "location": 25},
        "Sheet3": {"nameEn": 1, "nameAr": 2, "email": 3, "phone": 4, "organization": 5, "insurance": 8, "nationalId": 10, "preferredCenter": 24, "role": 20, "type": 21, "governorate": 22, "building": 24, "location": 25},
    }

    for sheet in workbook.worksheets:
        layout = layouts.get(sheet.title)
        if not layout:
            continue

        for row in sheet.iter_rows(values_only=True):
            values = [normalize_text(cell) for cell in row]
            if not any(values):
                continue

            maybe_append_person(
                people_rows,
                {
                    "sourceFile": path.name,
                    "sourceSheet": sheet.title,
                    "nameEn": values[layout["nameEn"]] if len(values) > layout["nameEn"] else "",
                    "nameAr": values[layout["nameAr"]] if len(values) > layout["nameAr"] else "",
                    "email": values[layout["email"]] if len(values) > layout["email"] else "",
                    "phone": values[layout["phone"]] if len(values) > layout["phone"] else "",
                    "organization": values[layout["organization"]] if len(values) > layout["organization"] else "",
                    "insuranceNumber": values[layout["insurance"]] if len(values) > layout["insurance"] else "",
                    "nationalId": values[layout["nationalId"]] if len(values) > layout["nationalId"] else "",
                    "preferredCenter": values[layout["preferredCenter"]] if len(values) > layout["preferredCenter"] else "",
                    "roleLabel": values[layout["role"]] if len(values) > layout["role"] else "",
                    "typeLabel": values[layout["type"]] if len(values) > layout["type"] else "",
                    "governorate": values[layout["governorate"]] if len(values) > layout["governorate"] else "",
                    "building": values[layout["building"]] if len(values) > layout["building"] else "",
                    "location": values[layout["location"]] if len(values) > layout["location"] else "",
                    "division": sheet.title,
                },
            )


def parse_final_roster(
    path: Path,
    people_rows: list[dict[str, Any]],
    final_rows: list[dict[str, Any]],
) -> None:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook["MAIN"]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = [normalize_text(cell) for cell in row]
        if not any(values):
            continue

        payload = {
            "sourceFile": path.name,
            "sourceSheet": sheet.title,
            "serial": values[0] if len(values) > 0 else "",
            "roomEst1": values[1] if len(values) > 1 else "",
            "roomEst2": values[2] if len(values) > 2 else "",
            "division": values[3] if len(values) > 3 else "",
            "nameEn": values[4] if len(values) > 4 else "",
            "nameAr": values[5] if len(values) > 5 else "",
            "email": values[6] if len(values) > 6 else "",
            "phone": values[7] if len(values) > 7 else "",
            "organization": values[8] if len(values) > 8 else "",
            "insuranceNumber": values[11] if len(values) > 11 else "",
            "nationalId": values[13] if len(values) > 13 else "",
            "preferredCenter": values[17] if len(values) > 17 else "",
            "roleLabel": values[23] if len(values) > 23 else "",
            "typeLabel": values[24] if len(values) > 24 else "",
            "governorate": values[25] if len(values) > 25 else "",
            "building": values[27] if len(values) > 27 else "",
            "location": values[28] if len(values) > 28 else "",
            "newOld": values[32] if len(values) > 32 else "",
        }

        final_rows.append(
            {
                "sourceFile": payload["sourceFile"],
                "sourceSheet": payload["sourceSheet"],
                "serial": payload["serial"],
                "roomEst1": clean_room_name(payload["roomEst1"]),
                "roomEst2": clean_room_name(payload["roomEst2"]),
                "division": payload["division"],
                "nameEn": payload["nameEn"],
                "nameAr": payload["nameAr"],
                "email": normalize_email(payload["email"]),
                "phone": normalize_phone(payload["phone"]),
                "organization": payload["organization"],
                "insuranceNumber": normalize_identifier(payload["insuranceNumber"]),
                "nationalId": normalize_identifier(payload["nationalId"]),
                "preferredCenter": payload["preferredCenter"],
                "roleLabel": payload["roleLabel"],
                "typeLabel": payload["typeLabel"],
                "governorate": payload["governorate"],
                "building": payload["building"],
                "location": payload["location"],
                "newOld": payload["newOld"],
            }
        )

        maybe_append_person(people_rows, payload)


def build_payload(base_dir: Path) -> dict[str, Any]:
    people_rows: list[dict[str, Any]] = []
    location_rows: list[dict[str, Any]] = []
    final_rows: list[dict[str, Any]] = []

    parse_aast_docx(
        base_dir / "1-AAST-EST-ALEX-Jan2026.docx",
        location_rows,
        "Alexandria",
        "Arab Academy Abu Qir",
        {
            "Building B": "Arab Academy Abu Qir Faculty of Engineering Building B",
            "Building C": "Arab Academy Abu Qir Faculty of Engineering Building C",
            "Building D": "Arab Academy Abu Qir Faculty of Engineering Building D",
            "Pharmacy Building": "Arab Academy Abu Qir Faculty of Pharmacy",
        },
    )
    parse_aast_docx(
        base_dir / "2-AAST-EST-HELIOPOLIS-Jan-2026(1).docx",
        location_rows,
        "Cairo",
        "Arab Academy Sheraton",
        {
            "Building A": "Arab Academy Sheraton Faculty of Engineering Building A",
            "Building B": "Arab Academy Sheraton Faculty of Engineering Building B",
            "PG Building": "Arab Academy Sheraton Faculty of Studies Post-Graduate Building",
            "L&M Building": "Arab Academy Sheraton Faculty of Language & Communication Building C",
            "CMT Building": "Arab Academy Sheraton Faculty of Management & Technology",
            "CITL Building": "Arab Academy Sheraton Faculty of International Transport & Logistics",
        },
    )
    parse_aast_docx(
        base_dir / "3-AAST-EST-SAMRT VILLAGE -Jan-2026(1).docx",
        location_rows,
        "Giza",
        "Arab Academy Smart Village",
        {
            "Building A": "Arab Academy Smart Village Building A",
            "Building B": "Arab Academy Smart Village Building B",
        },
    )
    parse_sadat_workbook(base_dir / "EST - Sadat City -Jan-2026.xlsx", location_rows)
    parse_future_workbook(base_dir / "Futur_EST_Jan-2026.xlsx", location_rows)
    parse_damietta_rooms(base_dir / "Jan-2026_Damitte.xls", location_rows)

    parse_ryada_people(base_dir / "Al Ryada- Jan 2026.xlsx", people_rows)
    parse_damietta_people(base_dir / "Damietta Staff 19-11-2025.xlsx", people_rows)
    parse_fue_people(base_dir / "FUE with Insurance Number and full names -2026.xlsx", people_rows)
    parse_uk_people(base_dir / "UK List.xlsx", people_rows)
    parse_final_roster(next(base_dir.glob("Final EST JAN 2026*.xlsm")), people_rows, final_rows)

    return {
        "baseDir": str(base_dir),
        "stats": {
            "peopleRows": len(people_rows),
            "locationRows": len(location_rows),
            "finalRosterRows": len(final_rows),
        },
        "peopleRows": people_rows,
        "locationRows": location_rows,
        "finalRosterRows": final_rows,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--base-dir", default=r"e:\est files")
    args = parser.parse_args()

    payload = build_payload(Path(args.base_dir))
    print(json.dumps(payload, ensure_ascii=False))


if __name__ == "__main__":
    main()
